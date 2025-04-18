import sqlite3
import os
from flask import Flask, request, jsonify, redirect, url_for, flash, render_template, session, make_response, Response
from werkzeug.utils import secure_filename
from PIL import Image
from functools import wraps
from datetime import timedelta, datetime
import csv
import io
from openpyxl import Workbook
from io import BytesIO
import pandas as pd
from flask import send_file
import platform
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.utils import ImageReader

app = Flask(__name__)
app.secret_key = 'kaon aron di ma brother'  # Secret key for session management

UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Update Flask configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=1)  # Increase session lifetime
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config['SESSION_COOKIE_SECURE'] = False  # Allow non-HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_TYPE'] = 'filesystem'

# Ensure the upload directory exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def get_db_connection():
    try:
        conn = sqlite3.connect('users.db')
        conn.row_factory = sqlite3.Row
        return conn
    except sqlite3.Error as e:
        print(f"Database connection error: {e}")
        return None

def resize_image(image_path, max_size=(300, 300)):
    with Image.open(image_path) as img:
        # Convert to RGB if image is in RGBA mode
        if img.mode == 'RGBA':
            img = img.convert('RGB')
            
        # Calculate aspect ratio
        aspect = img.width / img.height
        
        if img.width > max_size[0] or img.height > max_size[1]:
            if aspect > 1:
                # Width is greater than height
                new_width = min(img.width, max_size[0])
                new_height = int(new_width / aspect)
            else:
                # Height is greater than width
                new_height = min(img.height, max_size[1])
                new_width = int(new_height * aspect)
                
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
        # Save with high quality
        img.save(image_path, 'JPEG', quality=95, optimize=True)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            flash("Please log in first!", "warning")
            session.clear()  # Clear any invalid session data
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not all(session.get(key) for key in ['user_id', 'user_type', 'is_admin']):
            return redirect(url_for('login'))
        session.modified = True
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or 'user_type' not in session or session.get('user_type') != 'student':
            flash("Student access required!", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

@app.before_request
def validate_session():
    if request.endpoint and 'static' in request.endpoint:
        return  # Skip session validation for static files
        
    if 'user_id' in session:
        session.permanent = True
        session.modified = True
        
        # Keep admin session alive
        if session.get('user_type') == 'admin':
            session['is_admin'] = True
            session['last_activity'] = datetime.now().isoformat()

@app.route('/')
def home():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    
    if session.get('user_type') == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('dashboard'))

@app.route('/edit_profile', methods=['POST'])
@login_required
def edit_profile():
    student_id = session['user_id']
    firstname = request.form.get('firstname')
    lastname = request.form.get('lastname')
    course = request.form.get('course')
    year_level = request.form.get('year_level')
    email_address = request.form.get('email_address')

    # Handle Profile Picture Upload
    profile_picture = request.files.get('profile_picture')
    profile_picture_url = None

    if profile_picture and profile_picture.filename != '':
        # Create a unique filename using timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"profile_{student_id}_{timestamp}_{secure_filename(profile_picture.filename)}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        try:
            # Save the original file with proper extension
            file_extension = os.path.splitext(profile_picture.filename)[1].lower()
            if file_extension not in ['.jpg', '.jpeg', '.png']:
                raise ValueError("Invalid file format. Only JPG and PNG are allowed.")
                
            profile_picture.save(file_path)
            # Resize the image with improved quality
            resize_image(file_path)
            # Set the relative path for database storage
            profile_picture_url = os.path.join('uploads', filename).replace('\\', '/')
        except Exception as e:
            flash(f"Error saving profile picture: {str(e)}", "danger")
            return redirect(url_for('dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        if profile_picture_url:
            # Get current profile picture
            cursor.execute("SELECT profile_picture FROM students WHERE idno = ?", (student_id,))
            old_picture = cursor.fetchone()['profile_picture']

            # Delete old profile picture if it exists
            if old_picture:
                old_picture_path = os.path.join('static', old_picture)
                if os.path.exists(old_picture_path):
                    os.remove(old_picture_path)

            cursor.execute("""
                UPDATE students 
                SET firstname = ?, lastname = ?, course = ?, year_level = ?, 
                    email_address = ?, profile_picture = ?
                WHERE idno = ?
            """, (firstname, lastname, course, year_level, email_address, profile_picture_url, student_id))
        else:
            cursor.execute("""
                UPDATE students 
                SET firstname = ?, lastname = ?, course = ?, year_level = ?, 
                    email_address = ?
                WHERE idno = ?
            """, (firstname, lastname, course, year_level, email_address, student_id))

        conn.commit()
        flash("Profile updated successfully!", "success")
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
    finally:
        conn.close()

    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Clear any existing session
    if session.get('user_id'):
        session.clear()

    if request.method == 'POST':
        user_id = request.form.get('user_id')
        password = request.form.get('password')

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Check admin first
            cursor.execute("SELECT * FROM admin WHERE username = ? AND password = ?", (user_id, password))
            admin = cursor.fetchone()

            if admin:
                session.permanent = True  # Enable session expiry
                session['user_id'] = str(admin["id"])
                session['user_type'] = 'admin'
                session['username'] = admin["username"]
                session['is_admin'] = True  # Add explicit admin flag
                session.modified = True  # Ensure session is saved
                flash("Admin login successful!", "success")
                return redirect(url_for('admin_dashboard'))

            # Then check students
            cursor.execute("SELECT * FROM students WHERE idno = ? AND password = ?", (user_id, password))
            student = cursor.fetchone()

            if student:
                session.permanent = True  # Enable session expiry
                session['user_id'] = str(student["idno"])
                session['user_type'] = 'student'
                session['firstname'] = student["firstname"]
                session['lastname'] = student["lastname"]
                session['session_count'] = student["session_count"]
                flash("Login successful!", "success")
                return redirect(url_for('dashboard'))

            flash("Invalid credentials!", "danger")
            return redirect(url_for('login'))

        except sqlite3.Error as e:
            flash(f"Database error: {str(e)}", "danger")
            return redirect(url_for('login'))
        finally:
            conn.close()

    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        lastname = request.form.get('lastname')
        firstname = request.form.get('firstname')
        middlename = request.form.get('middlename', '')
        course = request.form.get('course')
        year_level = request.form.get('year_level')
        email = request.form.get('email')
        username = request.form.get('username')  
        password = request.form.get('password')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if student ID or username already exists
        cursor.execute("SELECT * FROM students WHERE idno = ? OR username = ?", (student_id, username))
        existing_user = cursor.fetchone()

        if existing_user:
            flash("User already exists. Try a different ID or username.", "danger")
            conn.close()
            return redirect(url_for('register'))

        try:
            # Set session count based on course
            session_count = 30 if course in ['BSIT', 'BSCS'] else 15
            
            cursor.execute(
                'INSERT INTO students (idno, lastname, firstname, midname, course, year_level, email_address, username, password, session_count) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (student_id, lastname, firstname, middlename, course, year_level, email, username, password, session_count)
            )
            conn.commit()
            flash("Registration successful! You can now log in.", "success")
        except sqlite3.Error as e:
            flash(f"Database error: {str(e)}", "danger")
        finally:
            conn.close()

        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    if session.get('user_type') != 'student':
        session.clear()
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch user information
    cursor.execute("""
        SELECT idno, firstname, lastname, midname, course, year_level, email_address, profile_picture, session_count
        FROM students WHERE idno = ?
    """, (session['user_id'],))
    
    user = cursor.fetchone()

    # Fetch announcements
    cursor.execute("SELECT * FROM announcements ORDER BY created_at DESC")
    announcements = cursor.fetchall()
    conn.close()

    if not user:
        flash("User not found!", "danger")
        return redirect(url_for('logout'))

    # Update session data
    session['firstname'] = user['firstname']
    session['lastname'] = user['lastname']
    session['session_count'] = user['session_count']

    # Convert user data into a dictionary for Jinja template
    user_data = {
        "idno": user["idno"],
        "firstname": user["firstname"],
        "lastname": user["lastname"],
        "midname": user["midname"],
        "course": user["course"] or "N/A",  # Default if empty
        "year_level": user["year_level"] or "N/A",  # Default if empty
        "email_address": user["email_address"] or "N/A",  # Default if empty
        "profile_picture": user["profile_picture"] if user["profile_picture"] else None,  # None if not set
        "session_count": user["session_count"] if user["session_count"] else 0
    }

    return render_template('dashboard.html', user=user_data, announcements=announcements)

@app.route('/sit_in', methods=['POST'])
def sit_in():
    id_number = request.form.get('id_number')
    name = request.form.get('name')
    sit_purpose = request.form.get('sit_purpose')
    laboratory = request.form.get('laboratory')
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    current_date = datetime.now().strftime('%Y-%m-%d')

    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Check existing active session
            cursor.execute("""
                SELECT * FROM current_sit_in 
                WHERE id_number = ? AND status = 'Active'
            """, (id_number,))
            existing_sit_in = cursor.fetchone()

            if existing_sit_in:
                flash("Student is already logged in for a sit-in session.", "danger")
                return redirect(url_for('admin_dashboard'))

            # Check session count
            cursor.execute("SELECT session_count FROM students WHERE idno = ?", (id_number,))
            session_count = cursor.fetchone()["session_count"]
            
            if session_count <= 0:
                flash("Student has no remaining session time.", "danger")
                return redirect(url_for('admin_dashboard'))

            # First insert into sit_in_records
            cursor.execute("""
                INSERT INTO sit_in_records 
                (id_number, name, purpose, lab, login_time, date)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (id_number, name, sit_purpose, laboratory, current_time, current_date))

            # Then insert into current_sit_in
            cursor.execute("""
                INSERT INTO current_sit_in 
                (id_number, name, purpose, sit_lab, session, status)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (id_number, name, sit_purpose, laboratory, session_count, 'Active'))

            conn.commit()
            flash("Sit-in logged successfully!", "success")
            return redirect(url_for('admin_dashboard'))
        
    except sqlite3.Error as e:
        print(f"Database error: {e}")  # Add debug print
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))

@app.route('/update_logout', methods=['POST'])
def update_logout():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get current sit-in details
        cursor.execute("""
            SELECT name, purpose, sit_lab FROM current_sit_in 
            WHERE id_number = ? AND status = 'Active'
        """, (data['id_number'],))
        current_session = cursor.fetchone()

        if current_session:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            current_date = datetime.now().strftime('%Y-%m-%d')
            
            # Insert into sit_in_history with all required fields
            cursor.execute("""
                INSERT INTO sit_in_history 
                (id_number, name, sit_purpose, laboratory, login_time, logout_time, date, action)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data['id_number'],
                current_session['name'],
                current_session['purpose'],
                current_session['sit_lab'],
                data.get('login_time', current_time),
                current_time,
                current_date,
                'Completed'
            ))
            
            # Update current_sit_in status
            cursor.execute("""
                UPDATE current_sit_in 
                SET status = 'Completed' 
                WHERE id_number = ? AND status = 'Active'
            """, (data['id_number'],))
            
            # Update session count
            cursor.execute("""
                UPDATE students 
                SET session_count = MAX(session_count - 1, 0) 
                WHERE idno = ?
            """, (data['id_number'],))
            
            conn.commit()
            return jsonify({"message": "Logout successful and history recorded"}), 200
        else:
            return jsonify({"message": "No active session found"}), 404
            
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"message": f"Database error: {str(e)}"}), 500
    finally:
        conn.close()

@app.route('/sit_in_history')
@login_required
def sit_in_history():
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Query the history with all fields
        cursor.execute("""
            SELECT 
                id,
                id_number,
                name,
                sit_purpose,
                laboratory,
                strftime('%Y-%m-%d %H:%M', login_time) as login_time,
                strftime('%Y-%m-%d %H:%M', logout_time) as logout_time,
                date,
                action,
                feedback,
                feedback_date
            FROM sit_in_history
            WHERE id_number = ?
            ORDER BY date DESC, login_time DESC
        """, (user_id,))
        
        history = cursor.fetchall()
        
        # Debug print
        print(f"Found {len(history)} records for user {user_id}")
        
        return render_template('sit_in_history.html', history=history)
        
    except sqlite3.Error as e:
        print(f"Database error in sit_in_history: {e}")
        flash("Error loading history", "danger")
        return redirect(url_for('dashboard'))
    finally:
        conn.close()

@app.route('/submit_feedback', methods=['POST'])
@login_required
def submit_feedback():
    try:
        history_id = request.form.get('history_id')
        feedback = request.form.get('feedback')
        user_id = session['user_id']
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if not history_id or not feedback:
            flash("Missing required information", "danger")
            return redirect(url_for('sit_in_history'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Update feedback in sit_in_history table
        cursor.execute("""
            UPDATE sit_in_history 
            SET feedback = ?,
                feedback_date = ?
            WHERE id = ? AND id_number = ?
        """, (feedback, current_time, history_id, user_id))
        
        conn.commit()
        flash("Thank you for your feedback!", "success")
            
    except sqlite3.Error as e:
        print(f"Database error: {str(e)}")
        flash(f"Error submitting feedback. Please try again.", "danger")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()
    
    return redirect(url_for('sit_in_history'))

@app.route('/reserve', methods=["GET", "POST"])
@login_required
def reserve():
    user_id = session['user_id']  # Get the logged-in user ID

    with get_db_connection() as conn:
        cursor = conn.cursor()

        # Fetch student details including session_count
        cursor.execute("SELECT idno, firstname, lastname, session_count FROM students WHERE idno = ?", (user_id,))
        user = cursor.fetchone()

        if not user:
            flash("User not found!", "danger")
            return redirect(url_for('logout'))

        remaining_session = user["session_count"]  # Sync remaining session with database

        if request.method == 'POST':
            purpose = request.form['purpose']
            lab = request.form['lab']
            time_in = request.form['time_in']
            date = request.form['date']

            # Prevent reservations if session count is zero
            if remaining_session <= 0:
                flash("You have no remaining session.", "danger")
                return redirect(url_for('reserve'))

            try:
                # Insert reservation without deducting session count
                cursor.execute('INSERT INTO reservations (id_number, student_name, sit_purpose, laboratory, time_in, date, remaining_sessions) VALUES (?, ?, ?, ?, ?, ?, ?)',
                               (user["idno"], f"{user['firstname']} {user['lastname']}", purpose, lab, time_in, date, remaining_session))

                conn.commit()

                flash("Reservation created successfully!", "success")
            except sqlite3.Error as e:
                flash(f"Database error: {str(e)}", "danger")

            return redirect(url_for('reserve'))

    # Pass updated remaining session count to template
    return render_template('reservation.html', 
                           id_number=user["idno"],
                           student_name=f"{user['firstname']} {user['lastname']}",
                           remaining_session=session["session_count"])

@app.route('/admin_dashboard', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_dashboard():
    session.modified = True  # Ensure session is refreshed
    
    if not session.get('is_admin'):
        session['is_admin'] = True
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        try:
            cursor.execute(
                'INSERT INTO announcements (title, content, created_at) VALUES (?, ?, ?)',
                (title, content, created_at)
            )
            conn.commit()
            flash("Announcement created successfully!", "success")
        except sqlite3.Error as e:
            flash(f"Database error: {str(e)}", "danger")

    # Fetch statistics
    cursor.execute("SELECT COUNT(*) FROM students")
    student_registered = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM current_sit_in WHERE status = 'Active'")
    current_sit_in = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM current_sit_in")
    total_sit_in = cursor.fetchone()[0]

    # Fetch announcements
    cursor.execute("SELECT id, title, content, created_at FROM announcements ORDER BY created_at DESC")
    announcements = cursor.fetchall()
    
    conn.close()
    
    return render_template('admin/admin_dashboard.html', student_registered=student_registered, current_sit_in=current_sit_in, total_sit_in=total_sit_in, announcements=announcements)

@app.route('/student_list')
@admin_required
def student_list():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Fetch students data
    cursor.execute("SELECT * FROM students")
    students = cursor.fetchall()
    
    conn.close()
    
    return render_template('admin/student_list.html', students=students)

@app.route('/get_student/<student_id>')
def get_student(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT idno, firstname, midname, lastname, session_count FROM students WHERE idno = ?", (student_id,))
    student = cursor.fetchone()
    conn.close()

    if student:
        student_data = {
            "success": True,
            "idno": student["idno"],
            "name": f"{student['firstname']} {student['midname']} {student['lastname']}",
            "session_count": student["session_count"]
        }
        return jsonify(student_data)
    else:
        return jsonify({"success": False})

@app.route('/current_sit_in')
@admin_required
def current_sit_in():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT s.idno, s.firstname, s.midname, s.lastname, s.course, s.year_level, 
               s.session_count, c.purpose, c.sit_lab, c.status
        FROM students s
        INNER JOIN current_sit_in c ON s.idno = c.id_number
        WHERE c.status = 'Active'
        ORDER BY c.sit_id_number DESC
    """)
    sit_ins = cursor.fetchall()
    conn.close()
    # Add debug print to check what's being returned
    print("Current sit-ins:", [dict(sit_in) for sit_in in sit_ins])
    return render_template('admin/current_sit_in.html', sit_ins=sit_ins)

@app.route('/end_session/<idno>', methods=['POST'])
@admin_required
def end_session(idno):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # First get the active session details from sit_in_records
            cursor.execute("""
                SELECT id_number, name, purpose, lab, login_time, date
                FROM sit_in_records 
                WHERE id_number = ? AND logout_time IS NULL
            """, (idno,))
            active_record = cursor.fetchone()

            if active_record:
                # Update sit_in_records with logout time
                cursor.execute("""
                    UPDATE sit_in_records 
                    SET logout_time = ? 
                    WHERE id_number = ? AND logout_time IS NULL
                """, (current_time, idno))

                # Insert into sit_in_history
                cursor.execute("""
                    INSERT INTO sit_in_history 
                    (id_number, name, sit_purpose, laboratory, login_time, logout_time, date, action)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    active_record['id_number'],
                    active_record['name'],
                    active_record['purpose'],
                    active_record['lab'],
                    active_record['login_time'],
                    current_time,
                    active_record['date'],
                    'Completed'
                ))

                # Update current_sit_in status
                cursor.execute("""
                    UPDATE current_sit_in 
                    SET status = 'Completed' 
                    WHERE id_number = ? AND status = 'Active'
                """, (idno,))

                # Update session count
                cursor.execute("""
                    UPDATE students 
                    SET session_count = MAX(session_count - 1, 0) 
                    WHERE idno = ?
                """, (idno,))

                conn.commit()
                return jsonify({"message": "Session ended successfully"}), 200
            else:
                return jsonify({"message": "No active session found"}), 404

    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"message": f"Database error: {str(e)}"}), 500
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": "An error occurred"}), 500

@app.route('/announcement/edit/<int:id>', methods=['POST'])
@admin_required
def edit_announcement(id):
    title = request.form.get('title')
    content = request.form.get('content')

    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(
                'UPDATE announcements SET title = ?, content = ? WHERE id = ?',
                (title, content, id)
            )
            conn.commit()
            flash("Announcement updated successfully!", "success")
        except sqlite3.Error as e:
            flash(f"Database error: {str(e)}", "danger")
        finally:
            cursor.close()

    return redirect(url_for('admin_dashboard'))

@app.route('/announcement/delete/<int:id>', methods=['POST'])
@admin_required
def delete_announcement(id):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute('DELETE FROM announcements WHERE id = ?', (id,))
            conn.commit()
            flash("Announcement deleted successfully!", "success")
        except sqlite3.Error as e:
            flash(f"Database error: {str(e)}", "danger")
        finally:
            cursor.close()

    return redirect(url_for('admin_dashboard'))

@app.route('/delete_student/<student_id>', methods=['POST'])
@admin_required
def delete_student(student_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Delete student's profile picture if exists
        cursor.execute("SELECT profile_picture FROM students WHERE idno = ?", (student_id,))
        student = cursor.fetchone()
        if student and student['profile_picture']:
            profile_pic_path = os.path.join('static', student['profile_picture'])
            if os.path.exists(profile_pic_path):
                os.remove(profile_pic_path)

        # Delete student's records from related tables
        cursor.execute("DELETE FROM sit_in_records WHERE id_number = ?", (student_id,))
        cursor.execute("DELETE FROM current_sit_in WHERE id_number = ?", (student_id,))
        cursor.execute("DELETE FROM sit_in_history WHERE id_number = ?", (student_id,))
        cursor.execute("DELETE FROM students WHERE idno = ?", (student_id,))
        
        conn.commit()
        flash("Student removed successfully", "success")
    except sqlite3.Error as e:
        flash(f"Error removing student: {str(e)}", "danger")
    finally:
        conn.close()
    
    return redirect(url_for('student_list'))

@app.route('/logout')
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for('login'))

@app.route('/sit_in_purposes')
@admin_required
def sit_in_purposes():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT sit_purpose, COUNT(*) as count
        FROM sit_in_history
        GROUP BY sit_purpose
    """)
    results = cursor.fetchall()
    conn.close()

    labels = [row['sit_purpose'] for row in results]
    counts = [row['count'] for row in results]

    return jsonify({"labels": labels, "counts": counts})

@app.route('/sit_in_records')
@admin_required
def sit_in_records():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                id_number,
                name,
                purpose,
                lab,
                strftime('%Y-%m-%d %H:%M', login_time) as login_time,
                CASE 
                    WHEN logout_time IS NOT NULL THEN strftime('%Y-%m-%d %H:%M', logout_time)
                    ELSE 'Active'
                END as logout_time,
                strftime('%Y-%m-%d', date) as date,
                CASE 
                    WHEN login_time IS NOT NULL AND logout_time IS NOT NULL
                    THEN ROUND((julianday(logout_time) - julianday(login_time)) * 24, 1)
                    ELSE NULL
                END as duration
            FROM sit_in_records
            ORDER BY 
                CASE WHEN logout_time IS NULL THEN 0 ELSE 1 END,
                datetime(login_time) DESC
        """)
        records = cursor.fetchall()
        
        return render_template('admin/sit_in_records.html', records=records)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))
    finally:
        if conn:
            conn.close()

@app.route('/reports')
@admin_required
def reports():
    try:
        lab = request.args.get('lab', '')
        purpose = request.args.get('purpose', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT 
                id_number,
                name,
                purpose,
                lab,
                strftime('%Y-%m-%d %H:%M', login_time) as login_time,
                CASE 
                    WHEN logout_time IS NOT NULL THEN strftime('%Y-%m-%d %H:%M', logout_time)
                    ELSE 'Active'
                END as logout_time,
                strftime('%Y-%m-%d', date) as date,
                CASE 
                    WHEN login_time IS NOT NULL AND logout_time IS NOT NULL
                    THEN ROUND((julianday(logout_time) - julianday(login_time)) * 24, 1)
                    ELSE NULL
                END as duration
            FROM sit_in_records
            WHERE 1=1
        """
        params = []
        
        if lab:
            query += " AND lab = ?"
            params.append(lab)
        if purpose:
            query += " AND purpose = ?"
            params.append(purpose)
        
        query += " ORDER BY date DESC, login_time DESC"
        
        cursor.execute(query, params)
        reports = cursor.fetchall()
        
        # Get unique labs and purposes
        cursor.execute("SELECT DISTINCT lab FROM sit_in_records WHERE lab IS NOT NULL ORDER BY lab")
        labs = [row['lab'] for row in cursor.fetchall()]
        
        cursor.execute("SELECT DISTINCT purpose FROM sit_in_records WHERE purpose IS NOT NULL ORDER BY purpose")
        purposes = [row['purpose'] for row in cursor.fetchall()]

        if request.headers.get('Accept') == 'application/json':
            return jsonify({
                'reports': [dict(row) for row in reports]
            })

        return render_template('admin/reports.html',
                             reports=reports,
                             labs=labs,
                             purposes=purposes,
                             selected_lab=lab,
                             selected_purpose=purpose)
                             
    except Exception as e:
        print(f"Error in reports: {str(e)}")
        flash("Error loading reports", "danger")
        return redirect(url_for('admin_dashboard'))
    finally:
        if conn:
            conn.close()

@app.route('/export_reports/<format>')
@admin_required
def export_reports(format):
    try:
        lab = request.args.get('lab', '')
        purpose = request.args.get('purpose', '')
        
        # Build query with filters
        query = """
            SELECT 
                id_number, name, purpose, lab,
                strftime('%Y-%m-%d %H:%M', login_time) as login_time,
                strftime('%Y-%m-%d %H:%M', logout_time) as logout_time,
                strftime('%Y-%m-%d', date) as date,
                CASE 
                    WHEN login_time IS NOT NULL AND logout_time IS NOT NULL
                    THEN ROUND((julianday(logout_time) - julianday(login_time)) * 24, 1)
                    ELSE NULL
                END as duration
            FROM sit_in_records
            WHERE 1=1
        """
        params = []
        
        if lab:
            query += " AND lab = ?"
            params.append(lab)
        if purpose:
            query += " AND purpose = ?"
            params.append(purpose)
            
        query += " ORDER BY date DESC, login_time DESC"
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params)
        reports = cursor.fetchall()
        conn.close()

        if not reports:
            flash("No data available for export", "warning")
            return redirect(url_for('reports'))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'Sit-in_Monitoring_System_Report{timestamp}'

        # Add filter info to filename
        if lab or purpose:
            filters = []
            if lab: filters.append(f"lab_{lab}")
            if purpose: filters.append(f"purpose_{purpose}")
            filename += f"_{'_'.join(filters)}"

        if format == 'csv':
            return export_to_csv(reports, filename)
        elif format == 'excel':
            return export_to_excel(reports, filename)
        elif format == 'pdf':
            return export_to_pdf(reports, filename, lab=lab, purpose=purpose)
        else:
            flash("Invalid export format", "error")
            return redirect(url_for('reports'))

    except Exception as e:
        flash(f"Export error: {str(e)}", "error")
        return redirect(url_for('reports'))

def export_to_csv(reports, filename):
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['ID Number', 'Name', 'Purpose', 'Laboratory', 
                        'Login Time', 'Logout Time', 'Date', 'Duration (hours)'])
        
        for report in reports:
            writer.writerow([
                report['id_number'],
                report['name'],
                report['purpose'],
                report['lab'],
                report['login_time'],
                report['logout_time'],
                report['date'],
                f"{report['duration']:.1f}" if report['duration'] else ''
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename={filename}.csv',
                'Content-Type': 'text/csv; charset=utf-8'
            }
        )
    except Exception as e:
        raise Exception(f"CSV export failed: {str(e)}")

def export_to_excel(reports, filename):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Computer Laboratory Sit-in Monitoring System Report"
        
        # Headers
        headers = ['ID Number', 'Name', 'Purpose', 'Laboratory', 
                  'Login Time', 'Logout Time', 'Date', 'Duration (hours)']
        ws.append(headers)
        
        # Data
        for report in reports:
            ws.append([
                report['id_number'],
                report['name'],
                report['purpose'],
                report['lab'],
                report['login_time'],
                report['logout_time'],
                report['date'],
                f"{report['duration']:.1f}" if report['duration'] else ''
            ])
        
        # Style the worksheet
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{filename}.xlsx'
        )
    except Exception as e:
        raise Exception(f"Excel export failed: {str(e)}")

def add_header_with_logos(canvas, doc):
    # Save canvas state
    canvas.saveState()

    # Define logo paths and sizes - slightly reduced
    left_logo_path = 'static/images/uc_logo.jpg'
    right_logo_path = 'static/images/css.png'
    logo_width = 65  # Reduced size
    logo_height = 65  # Reduced size

    # Check if logo files exist
    if os.path.exists(left_logo_path) and os.path.exists(right_logo_path):
        # Add left logo - adjusted position
        canvas.drawImage(
            left_logo_path,
            doc.leftMargin,
            doc.pagesize[1] - logo_height - 15,  # Adjusted position
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True
        )

        # Handle right logo with transparency
        try:
            right_logo = Image.open(right_logo_path)
            if right_logo.mode != 'RGBA':
                right_logo = right_logo.convert('RGBA')
            
            # Create a white background image
            white_bg = Image.new('RGBA', right_logo.size, 'WHITE')
            right_logo = Image.alpha_composite(white_bg, right_logo)
            right_logo = right_logo.convert('RGB')
            
            temp_path = 'static/images/temp_right_logo.jpg'
            right_logo.save(temp_path, 'JPEG', quality=95)
            
            # Draw right logo - adjusted position
            canvas.drawImage(
                temp_path,
                doc.pagesize[0] - doc.rightMargin - logo_width - 10,
                doc.pagesize[1] - logo_height - 15,  # Adjusted position
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True
            )
            
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
        except Exception as e:
            print(f"Error processing right logo: {e}")

    # Add title section with adjusted spacing
    canvas.setFont("Helvetica-Bold", 15)  # Slightly reduced font size
    canvas.setFillColor(colors.HexColor('#4A3599'))
    # Center the title text
    title = "COMPUTER LABORATORY SIT-IN MONITORING SYSTEM REPORT"
    title_width = canvas.stringWidth(title, "Helvetica-Bold", 15)
    canvas.drawString(
        (doc.pagesize[0] - title_width) / 2,
        doc.pagesize[1] - 45,  # Adjusted position
        title
    )

    # Add page number and timestamp with adjusted positions
    page_num = canvas.getPageNumber()
    text = f"Page {page_num}"
    canvas.setFont("Helvetica", 9)
    canvas.setFillColor(colors.grey)
    
    # Footer content
    canvas.drawRightString(
        doc.pagesize[0] - doc.rightMargin,
        doc.bottomMargin - 20,
        text
    )
    
    canvas.drawString(
        doc.leftMargin,
        doc.bottomMargin - 20,
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )

    canvas.restoreState()

def export_to_pdf(reports, filename, lab=None, purpose=None):
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=40,  # Reduced margins
            leftMargin=40,
            topMargin=110,  # Reduced top margin
            bottomMargin=40
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Reduced spacing after logos and title
        elements.append(Spacer(1, 20))  # Reduced from 40
        
        # Add report info with center alignment
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=11,  # Slightly reduced font size
            textColor=colors.grey,
            alignment=1,
            spaceAfter=10  # Reduced spacing
        )
        
        elements.append(Paragraph(
            f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", 
            header_style
        ))
        
        # Add filters if present with reduced spacing
        if lab or purpose:
            filter_text = []
            if lab: filter_text.append(f"Laboratory: {lab}")
            if purpose: filter_text.append(f"Purpose: {purpose}")
            elements.append(Paragraph(
                "Filters: " + ", ".join(filter_text), 
                header_style
            ))
        
        elements.append(Spacer(1, 10))  # Reduced spacing
        
        # Create summary statistics with compact styling
        total_duration = sum(r['duration'] or 0 for r in reports)
        avg_duration = total_duration / len(reports) if reports else 0
        
        summary_data = [
            ['Total Records:', str(len(reports))],
            ['Total Hours:', f"{total_duration:.1f}"],
            ['Average Duration:', f"{avg_duration:.1f} hours"]
        ]
        
        summary_table = Table(summary_data, colWidths=[90, 90])  # Reduced widths
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),  # Reduced font size
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),  # Reduced padding
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 10))  # Reduced spacing
        
        # Create main table data
        data = [['ID Number', 'Name', 'Purpose', 'Laboratory', 
                 'Login Time', 'Logout Time', 'Date', 'Duration']]
        
        for report in reports:
            data.append([
                report['id_number'],
                report['name'],
                report['purpose'],
                report['lab'],
                report['login_time'],
                report['logout_time'],
                report['date'],
                f"{report['duration']:.1f} hrs" if report['duration'] else 'N/A'
            ])
        
        # Create and style the main table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A3599')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),  # Increased font size
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),  # Increased font size
            ('BOTTOMPADDING', (0, 0), (-1, 0), 15),  # Increased padding
            ('TOPPADDING', (0, 0), (-1, -1), 8),    # Added top padding
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),  # Added bottom padding
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        # Calculate column widths based on content
        col_widths = ['10%', '20%', '15%', '12%', '13%', '13%', '10%', '7%']
        table._argW = [doc.width * float(width.strip('%'))/100 for width in col_widths]
        
        elements.append(table)
        
        # Build the PDF with new header function
        doc.build(
            elements,
            onFirstPage=add_header_with_logos,
            onLaterPages=add_header_with_logos
        )
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'{filename}.pdf'
        )
    except Exception as e:
        raise Exception(f"PDF export failed: {str(e)}")

@app.route('/feedback_reports')
@admin_required
def feedback_reports():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Enhanced query to get more detailed feedback information
        query = """
            SELECT 
                h.id,
                s.idno,
                s.firstname,
                s.lastname,
                s.course,
                s.year_level,
                h.sit_purpose,
                h.laboratory,
                h.feedback,
                h.date,
                h.login_time,
                h.logout_time,
                h.feedback_date,
                ROUND((julianday(h.logout_time) - julianday(h.login_time)) * 24, 1) as duration
            FROM sit_in_history h
            JOIN students s ON h.id_number = s.idno
            WHERE h.feedback IS NOT NULL 
            AND h.feedback != ''
            ORDER BY h.feedback_date DESC
        """
        
        cursor.execute(query)
        feedbacks = cursor.fetchall()
        
        conn.close()
        return render_template('admin/feedback_report.html',
                             feedbacks=feedbacks)
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))

@app.route('/reset_session/<student_id>', methods=['POST'])
@admin_required
def reset_session(student_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get student's course
        cursor.execute("SELECT course FROM students WHERE idno = ?", (student_id,))
        student = cursor.fetchone()
        
        if student:
            # Set session count based on course
            session_count = 30 if student['course'] in ['BSIT', 'BSCS'] else 15
            cursor.execute("UPDATE students SET session_count = ? WHERE idno = ?", (session_count, student_id))
            
        conn.commit()
        flash("Session count reset successfully", "success")
    except sqlite3.Error as e:
        flash(f"Error resetting session: {str(e)}", "danger")
    finally:
        conn.close()
    
    return redirect(url_for('student_list'))

@app.route('/reset_all_sessions', methods=['POST'])
@admin_required
def reset_all_sessions():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Reset session counts based on course
        cursor.execute("""
            UPDATE students 
            SET session_count = CASE 
                WHEN course IN ('BSIT', 'BSCS') THEN 30 
                ELSE 15 
            END
        """)
        
        conn.commit()
        flash("All session counts reset successfully", "success")
    except sqlite3.Error as e:
        flash(f"Error resetting sessions: {str(e)}", "danger")
    finally:
        conn.close()
    
    return redirect(url_for('student_list'))

@app.errorhandler(Exception)
def handle_error(error):
    print(f"Error occurred: {str(error)}")  # Add logging
    if 'user_id' in session:
        session.modified = True  # Try to preserve session
        return redirect(url_for('admin_dashboard' if session.get('is_admin') else 'dashboard'))
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
    # app.run(debug=True, host='192.168.254.199', port=5000)